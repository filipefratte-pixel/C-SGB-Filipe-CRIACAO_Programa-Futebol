import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Any, Tuple
import database

def parse_and_validate_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Lê um arquivo Excel (.xlsx ou .xls) e valida os cadastros de mensalistas.
    Retorna o resumo detalhado da importação.
    """
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    
    headers = []
    for cell in sheet[1]:
        val = str(cell.value or '').strip().upper()
        headers.append(val)
        
    # Check required headers
    if 'NOME' not in headers or 'ESTRELAS' not in headers:
        raise ValueError("O arquivo Excel deve conter as colunas obrigatórias 'NOME' e 'ESTRELAS'.")

    nome_col = headers.index('NOME')
    estrelas_col = headers.index('ESTRELAS')
    id_col = headers.index('ID') if 'ID' in headers else -1
    ativo_col = headers.index('ATIVO') if 'ATIVO' in headers else -1

    existing_mensalistas = database.get_all_mensalistas()
    existing_by_name = {m['nome'].strip().lower(): m for m in existing_mensalistas}
    existing_by_id = {m['id']: m for m in existing_mensalistas}

    novos = []
    atualizados = []
    duplicados = []
    invalidos = []
    
    seen_in_file = set()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue  # Skip completely empty rows
            
        raw_nome = row[nome_col] if nome_col < len(row) else None
        raw_estrelas = row[estrelas_col] if estrelas_col < len(row) else None
        raw_id = row[id_col] if id_col >= 0 and id_col < len(row) else None
        raw_ativo = row[ativo_col] if ativo_col >= 0 and ativo_col < len(row) else True

        # Validation: Nome
        if not raw_nome or not str(raw_nome).strip():
            invalidos.append({"linha": row_idx, "motivo": "Nome está em branco."})
            continue
            
        nome_str = str(raw_nome).strip()
        nome_lower = nome_str.lower()
        
        # Validation: Estrelas
        try:
            estrelas_val = int(float(str(raw_estrelas).strip()))
            if not (1 <= estrelas_val <= 5):
                invalidos.append({"linha": row_idx, "nome": nome_str, "motivo": f"Estrelas ({estrelas_val}) devem estar entre 1 e 5."})
                continue
        except (ValueError, TypeError):
            invalidos.append({"linha": row_idx, "nome": nome_str, "motivo": f"Valor inválido de estrelas ('{raw_estrelas}')."})
            continue

        # Parse Ativo
        ativo_bool = True
        if raw_ativo is not None:
            raw_ativo_str = str(raw_ativo).strip().lower()
            if raw_ativo_str in ('false', '0', 'não', 'nao', 'n', 'f', 'inativo'):
                ativo_bool = False
            elif raw_ativo_str in ('true', '1', 'sim', 's', 't', 'ativo'):
                ativo_bool = True

        # Duplicate check within the file
        if nome_lower in seen_in_file:
            duplicados.append({"linha": row_idx, "nome": nome_str, "motivo": "Nome duplicado dentro do próprio arquivo Excel."})
            continue
        seen_in_file.add(nome_lower)

        # Check against database
        existing = existing_by_name.get(nome_lower)
        if existing:
            # Player exists in DB -> Will update stars and active state if changed
            atualizados.append({
                "id": existing['id'],
                "nome": nome_str,
                "estrelas": estrelas_val,
                "ativo": ativo_bool,
                "estrelas_antigo": existing['estrelas']
            })
        else:
            novos.append({
                "nome": nome_str,
                "estrelas": estrelas_val,
                "ativo": ativo_bool
            })

    total_importado = len(novos) + len(atualizados)
    
    return {
        "novos": novos,
        "atualizados": atualizados,
        "duplicados": duplicados,
        "invalidos": invalidos,
        "total_importado": total_importado
    }

def confirm_excel_import(import_summary: Dict[str, Any]) -> Tuple[int, int]:
    """
    Aplica as inserções e atualizações validadas no banco de dados.
    """
    added_count = 0
    updated_count = 0
    
    for item in import_summary.get("novos", []):
        database.add_mensalista(item["nome"], item["estrelas"], item["ativo"])
        added_count += 1
        
    for item in import_summary.get("atualizados", []):
        database.update_mensalista(item["id"], item["nome"], item["estrelas"], item["ativo"])
        updated_count += 1
        
    return added_count, updated_count

def generate_excel_template() -> bytes:
    """
    Gera o modelo de planilha Excel (.xlsx) para cadastro de mensalistas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mensalistas"
    
    # Headers
    headers = ["ID", "NOME", "ESTRELAS", "ATIVO"]
    ws.append(headers)
    
    # Styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Sample data
    samples = [
        (1, "Carlos Silva", 5, "SIM"),
        (2, "Roberto Souza", 4, "SIM"),
        (3, "Lucas Mendes", 3, "SIM"),
        (4, "Felipe Andrade", 4, "SIM"),
        (5, "Gabriel Oliveira", 2, "NÃO")
    ]
    
    for row in samples:
        ws.append(list(row))
        
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def export_mensalistas_excel() -> bytes:
    """
    Exporta a lista atual de mensalistas cadastrados para um arquivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mensalistas Cadastrados"
    
    headers = ["ID", "NOME", "ESTRELAS", "ATIVO"]
    ws.append(headers)
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        
    mensalistas = database.get_all_mensalistas()
    for m in mensalistas:
        ws.append([m['id'], m['nome'], m['estrelas'], "SIM" if m['ativo'] else "NÃO"])
        
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
